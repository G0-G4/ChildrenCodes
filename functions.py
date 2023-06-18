import pickle
import PySimpleGUI as sg
import pandas as pd
import numpy as np
from openpyxl.utils.cell  import column_index_from_string, get_column_letter
from collections import defaultdict
from exceptions import *

CODE_COLUMN = 'код'
CLONE_COLUMN = 'клон'

def load_catalog(name):
    with open(name, 'rb') as f:
        cat = pickle.load(f)
    return cat

def set_columns(names, settings, check=True):
    res = {}
    layout = [
        [sg.T(name), sg.Input(get_column_letter(settings[name]+1) if name in settings else '', key=name)] for name in names
    ] + [[sg.B('ok')]]
    window =  sg.Window('setting up columns', layout, modal=True)
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED:
            break
        elif event == 'ok':
            for key in values:
                if values[key] == '':
                    continue
                try:
                    if check:
                        res[key] = column_index_from_string(values[key]) - 1
                    else:
                        res[key] = values[key] # for second tab no need for check
                except ValueError as e:
                    sg.popup(e)
                    break
            else:
                window.close()
                return res
    window.close()
    
def generate_children(cat):
    clones = defaultdict(list)
    for code, clone in zip(cat[CODE_COLUMN], cat[CLONE_COLUMN]):
        if not str(clone).isdigit():
            print('skipped clone ', clone)
            continue
        clone = int(clone)
        if clone != 0 and clone != code: # dont add mother code to its children
            clones[clone].append(code)
    return clones

def get_values(cat, codes, names):
    df = cat[cat[CODE_COLUMN].isin(codes)][names]
    return [list(df[name]) for name in names]

def get_range(c1: int, r1: int, c2: int, r2: int):
    c1, c2 = get_column_letter(c1), get_column_letter(c2)
    return f'{c1}{r1}:{c2}{r2}'

def insert_values(ws, start, end, columns, values):
    for i, row in enumerate(ws.iter_rows(min_row = start, max_row = end)):
        for j, col in enumerate(columns):
            val = values[j][i]
            if type(val) == str:
                val = val.strip()
            print('      inserting', val, 'into', (start+i, col+1))
            try:
                row[col].value = val
            except IndexError:
                raise InsertError(f'cant insert into {(start+i, col+1)}',)
            
def get_all_codes(ws, start, settings):
    # save codes indexes
    codes = dict()
    i = start
    code_id = settings[CODE_COLUMN]
    while True:
        row = ws[i]
        if row[code_id].value == None:
            break
        try:
            codes[int(row[code_id].value)] = i
            # codes.add(int(row[code_id].value))
        except Exception as e:
            print(f'Wrong mother code {row[code_id].value}')
            # raise MotherCodeError(f'Wrong mother code {row[code_id].value}')
        i += 1
    return codes

def get_mother(code, cat):
    if len(cat[cat[CODE_COLUMN] == code][CLONE_COLUMN]) != 1:
        return 0
    return int(cat[cat[CODE_COLUMN] == code][CLONE_COLUMN])

# BIAS = 0
# def add_children(ws, start, children, settings, cat):
#     for every code
#     if code is not mother, and mother not in file add mother above
#     if mother in file, move it above
#     if code is mother do nothing and go to adding children
#     for this mother code add children from catalog
#     for this mother code add children from file
#     global BIAS
#     BIAS = 0
#     if CODE_COLUMN not in settings:
#         raise CodeColumnNotFound('please provide column for mother code')
#     all_codes = get_all_codes(ws, start, settings)
#     code_id = settings[CODE_COLUMN]
#     if code_id >= len(ws[1]):
#         raise OutOfBounds('provided mother code column is not in the table')
#     i = start
#     codes = all_codes.copy() # codes used to move rows
#     while True:
#         row = ws[i]
#         if row[code_id].value == None:
#             break
#         try:
#             code = int(row[code_id].value)
#         except Exception as e:
#             print(f'Wrong mother code {row[code_id].value}')
#             # raise MotherCodeError(f'Wrong mother code {row[code_id].value}')
        
#         mother = get_mother(code, cat)
#         if code not in children and mother != 0:
#             if mother not in all_codes:
#                 add_mother_from_catalog()
#             else:
#                 move_mother()
#         elif code in children:
#             move_children()
#             add_children()
#             print(code, 'inserting mother', mother)
#             if i != ws.max_row:
#                 rang = get_range(1, i, ws.max_column, ws.max_row)
#                 print(rang, ' moving down 1')
#                 ws.move_range(rang, rows = 1, translate = True) # moving instead of inserting for not managing formulas
#             values = get_values(cat, [mother], settings.keys())
#             insert_values(ws, i, i, settings.values(), values)
#             all_codes[mother] = None
#             i += 1
#             BIAS += 1
#             print('moved to ', i, 'code = ', code, f'{BIAS=}')
#             continue
#         if code in children:
#             not_used_children = list(filter(lambda x: x not in all_codes, children[code])) # add not used children from catalog
#             num = len(not_used_children)
#             print(code, 'moving down', len(not_used_children), 'rows')
#             # ws.insert_rows(i+1, num) # insert_rows inserts before index, indexing starts from one there for add 1
#             if i != ws.max_row:
#                 rang = get_range(1, i+1, ws.max_column, ws.max_row)
#                 print(rang)
#                 ws.move_range(rang, rows = num, translate = True) # moving instead of inserting for not managing formulas
#             values = get_values(cat, not_used_children, settings.keys())
#             insert_values(ws, i+1, i+num, settings.values(), values)
#             i += num
#         i += group_codes(ws, code, i, codes, children, cat)
#         i += 1
#     print('=============')


def group_codes(ws, code, i, all_codes, children, cat):
    global BIAS
    if len(cat[cat[CODE_COLUMN] == code][CLONE_COLUMN]) != 1:
        return 0
    mother = int(cat[cat[CODE_COLUMN] == code][CLONE_COLUMN])
    if mother == 0:
        return 0
    print(f'{code=}, {mother=}')
    brothers = list(filter(lambda x: x in all_codes and x != code, children[mother])) # all codes that are in file
    rang = get_range(1, i+1, ws.max_column, ws.max_row)
    print(f'moving {rang} {len(brothers)}')
    ws.move_range(rang, rows = len(brothers), translate = True) # free space for them
    BIAS += len(brothers)
    print(f'{BIAS=}')
    for c in brothers:
        print(f'{c=}, {all_codes[c]=}')
        ii = all_codes[c] + BIAS
        rang = get_range(1, ii, ws.max_column, ii)
        del all_codes[c] # need to move only for first code from group
        print(f'{i=}, {ii=}, moving {rang} {i - ii + 1}')
        ws.move_range(rang, rows = i - ii, translate = True) # move under current location
        ws.delete_rows(ii)
        rang = get_range(1, ii+1, ws.max_column, ws.max_row)
        print(f'moving {rang} -1')
        ws.move_range(rang, rows=-1, translate=True)
        BIAS -= 1
        print(f'{BIAS=}')
        i += 1
    return len(brothers)




# def mother_children_table(cat, codes, settings):
#     cat_grouped = cat.set_index([CLONE_COLUMN, CODE_COLUMN])
#     cat_grouped.sort_index(inplace = True)
#     print(settings.keys())
#     return cat_grouped.loc[codes]

# def read_codes(file, settings):
#     return pd.read_excel(file).iloc[:, settings[CODE_COLUMN]]

def check_df(df):
    if df.columns[0]!= CODE_COLUMN:
        return False
    return True

def group(cat, clones, copy, df):
    add_children = []
    # searching childten to add
    for _, row in df.iterrows():
        if row[CODE_COLUMN] in clones:
            for ch in clones[row[CODE_COLUMN]]:
                if ch not in df[CODE_COLUMN].values:
                    add_children.append(ch)
    # merging
    df = pd.concat([df, (pd.DataFrame({CODE_COLUMN: add_children}))])
    res = df.merge(cat, on=[CODE_COLUMN], how='left', suffixes=('','catalog'))
    res = res.set_index([CLONE_COLUMN, CODE_COLUMN])
    res.sort_index(inplace = True)
    # copying values
    if copy:
        for i in res.index.get_level_values(0):
            n = len(res.loc[i])
            for name in df.columns[1:]:
                res.loc[i, name] = np.full(n, res.loc[i].iloc[0][name])
    return res

def check_box_window(initial):
    layout = [
        [sg.Column([[sg.Checkbox(name, key=name, default=val)] for name, val in initial.items()], scrollable=True, size=(200, 300)), sg.B('ok')]
]
    window =  sg.Window('choose', layout, modal=True)
    while True:
        event, values = window.read()
        if event:
            window.close()
        if event == sg.WIN_CLOSED:
            return []
        if event == 'ok':
            return list(map(lambda x: x[0], filter(lambda x: x[1], values.items())))

if __name__ == '__main__':
    print(check_box_window({'a':1, 'b':0}))