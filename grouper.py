from exceptions import *
from openpyxl.utils.cell  import column_index_from_string, get_column_letter
from collections import defaultdict
from functions import CODE_COLUMN, CLONE_COLUMN

class Grouper():
    def __init__(self, wb, start, children, columns, cat):
        print(columns)
        self.wb = wb
        self.ws = wb.active
        self.i = start - 1 # т.к потом будет move который сдвинет
        self.children = children
        self.columns = columns
        self.cat = cat
        if CODE_COLUMN not in columns:
            raise Exception('code column is not provided')
        self.code_id = columns[CODE_COLUMN]
        self.all_codes = self.get_all_codes()
        self.catalog_codes = set(cat[CODE_COLUMN])
        if self.code_id >= len(self.ws[1]):
            raise OutOfBounds('provided mother code column is not in the table')
        
    def save(self, name):
        self.wb.save(name)

    def get_mother(self, code):
        if len(self.cat[self.cat[CODE_COLUMN] == code][CLONE_COLUMN]) != 1:
            return 0
        mother = self.cat[self.cat[CODE_COLUMN] == code][CLONE_COLUMN]
        try:
            mother = int(mother)
            return mother
        except:
            print('cant convert into int ', mother)
            return 0
    
    def is_mother(self, code):
        return code == self.get_mother(code)
    
    def move(self):
        self.i += 1
        row = self.ws[self.i]
        print('current i = ', self.i)
        if row[self.code_id].value == None:
            return False
        try:
            code = int(row[self.code_id].value)
            print('code = ', code)
        except Exception as e:
            print(f'Wrong mother code {row[self.code_id].value}')
        else:
            return code
        return -1
    
    def get_range(self, c1: int, r1: int, c2: int, r2: int):
        c1, c2 = get_column_letter(c1), get_column_letter(c2)
        return f'{c1}{r1}:{c2}{r2}'
    
    def update_rows(self, rows, frm, to):
        for code, i in self.all_codes.items():
            if frm <= i <= to:
                self.all_codes[code] += rows
    
    def move_range(self, rows, frm, to = None):
        if not to:
            to = self.ws.max_row
        rang = self.get_range(1, frm, self.ws.max_column, to)
        self.update_rows(rows, frm, to)
        print('moving ', rang, 'by ', rows)
        self.ws.move_range(rang, rows)

    def get_biased_index(self, code):
        return self.all_codes[code]
    
    def move_row(self, row, rows, code):
        self.move_range(rows=rows, frm=row, to=row)
        self.move_range(rows=-1, frm=row+1)
    
    def move_mother(self, mother):
        print('moving mother')
        self.move_range(rows=1, frm=self.i)
        mother_id = self.get_biased_index(mother)
        self.move_row(mother_id, self.i - mother_id, mother)
        print('staying at mother ', self.i)

    def get_values(self, codes, names):
        df = self.cat[self.cat[CODE_COLUMN].isin(codes)][names]
        values = [list(df[name]) for name in names]
        if len(codes) != len(values[0]):
            print('some values were not found')
            return values
        return values

    def add_mother(self, mother):
        print('adding mother ', mother)
        if mother not in self.catalog_codes:
            print('mother not found in catalog')
            values = [[mother]] + [['мат кода нет в каталоге'] for i in range(len(self.columns.keys()) - 1)]
        else:
            values = self.get_values([mother], self.columns.keys())
        self.move_range(rows=1, frm=self.i)
        print('inserting mother into ', self.i)
        self.insert_values(self.i, self.i, self.columns.values(), values)
        print('staying at mother ', self.i)

    def move_children(self, mother):
        print('moving children')
        children = list(filter(lambda x: x in self.all_codes, self.children[mother]))
        for child in children:
            ii = self.get_biased_index(child)
            self.move_row(row = ii, rows = self.i - ii + 1, code=child)
            self.move()

    def add_children(self, mother):
        print('adding children')
        children = list(filter(lambda x: x not in self.all_codes, self.children[mother]))
        values = self.get_values(children, self.columns.keys())
        self.insert_values(self.i+1, self.i+len(children), self.columns.values(), values)
        self.i += len(children) # point to the last inserted values

    def insert_values(self, start, end, columns, values):
        print('inserting')
        print(values)
        for j, col in enumerate(columns):
            for i, row in enumerate(self.ws.iter_rows(min_row = start, max_row = end)):
                if i >= len(values[j]):
                    val = 'не найден'
                else:
                    val = values[j][i]
                print('      inserting', val, 'into', (start+i, col+1))
                if type(val) == str:
                    val = val.strip()
                try:
                    row[col].value = val
                except IndexError:
                    raise InsertError(f'cant insert into {(start+i, col+1)}',)

    def process(self):
        while code := self.move():
            mother = code
            if not self.is_mother(code):
                mother = self.get_mother(code)
                if mother in self.all_codes:
                    self.move_mother(mother)
                elif mother != 0:
                    self.add_mother(mother)
            if mother == 0:
                continue
            self.move_range(rows = len(self.children[mother]), frm=self.i + 1) # free space for all brothers
            self.move_children(mother)
            self.add_children(mother)

    def get_all_codes(self):
        print('getting all codes')
        i = self.i
        codes = dict()
        while code := self.move():
            codes[code] = self.i
        self.i = i
        print('=================')
        return codes

    
